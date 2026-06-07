from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import datetime


def generate_excel(data, workbook_path=None):

    print(
        "WORKBOOK PATH RECEIVED:",
        workbook_path
    )

    if workbook_path:

        print(
            "LOADING EXISTING WORKBOOK"
        )

        wb = load_workbook(
            workbook_path
        )

    else:

        print(
            "CREATING NEW WORKBOOK"
        )

        wb = Workbook()

        months = [
            "January",
            "February",
            "March",
            "April",
            "May",
            "June",
            "July",
            "August",
            "September",
            "October",
            "November",
            "December"
        ]

        # Rename first sheet
        wb.active.title = months[0]

        # Create remaining sheets
        for month_name in months[1:]:
            wb.create_sheet(
                title=month_name
            )

    selected_month = data.get(
        "month",
        "January"
    )

    ws = wb[selected_month]

    # ==========================
    # STYLES
    # ==========================

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    title_font = Font(
        bold=True,
        size=16
    )

    header_font = Font(
        bold=True,
        size=11
    )

    center_align = Alignment(
        horizontal="center",
        vertical="center"
    )

    # ==========================
    # HEADER DATA
    # ==========================

    business_name = data.get("business_name", "")
    gstin = data.get("gstin", "")
    uid = data.get("uid", "")
    password = data.get("password", "")
    month = data.get("month", "")

    invoices = data.get("invoices", [])

    # ==========================
    # TOP HEADER
    # ==========================

    ws.merge_cells("A1:N1")
    ws["A1"] = business_name
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws.merge_cells("A2:N2")
    ws["A2"] = f"PURCHASES & SALES FOR THE MONTH OF {month}"
    ws["A2"].font = header_font
    ws["A2"].alignment = center_align

    ws["A4"] = f"GSTIN : {gstin}"
    ws["D4"] = f"UID : {uid}"
    ws["G4"] = f"PASSWORD : {password}"

    # ==========================
    # COLUMN WIDTHS
    # ==========================

    widths = {
        "A": 12,
        "B": 35,
        "C": 15,
        "D": 22,
        "E": 18,
        "F": 15,
        "G": 10,
        "H": 15,
        "I": 15,
        "J": 15,
        "K": 15,
        "L": 15,
        "M": 15,
        "N": 18
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # ==========================
    # TABLE HEADERS
    # ==========================

    headers = [
        "S.No",
        "Party Name",
        "Place",
        "GSTIN",
        "Invoice No",
        "Date",
        "Tax Rate",
        "HSN Code",
        "Net Amount",
        "CGST",
        "SGST",
        "IGST",
        "Total GST",
        "Gross Amount"
    ]

    # ==========================
    # PURCHASE SECTION
    # ==========================

    purchase_start = 7

    ws.cell(purchase_start, 1).value = "PURCHASES"
    ws.cell(purchase_start, 1).font = header_font

    for col_num, header in enumerate(headers, start=1):

        cell = ws.cell(
            purchase_start + 1,
            col_num
        )

        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align

    purchase_row = purchase_start + 2
    purchase_serial = 1

    purchase_net_total = 0
    purchase_cgst_total = 0
    purchase_sgst_total = 0
    purchase_igst_total = 0
    purchase_total_gst = 0
    purchase_gross_total = 0

    for invoice in invoices:

        if invoice.get("invoice_type") != "Purchase":
            continue

        try:
            invoice_date = datetime.strptime(
                invoice.get("invoice_date", ""),
                "%Y-%m-%d"
            ).strftime("%d-%m-%Y")
        except:
            invoice_date = invoice.get(
                "invoice_date",
                ""
            )

        net_amount = float(
            invoice.get("net_amount", 0) or 0
        )

        cgst = float(
            invoice.get("cgst", 0) or 0
        )

        sgst = float(
            invoice.get("sgst", 0) or 0
        )

        igst = float(
            invoice.get("igst", 0) or 0
        )

        total_gst = cgst + sgst + igst
        gross_amount = net_amount + total_gst

        purchase_net_total += net_amount
        purchase_cgst_total += cgst
        purchase_sgst_total += sgst
        purchase_igst_total += igst
        purchase_total_gst += total_gst
        purchase_gross_total += gross_amount

        row_data = [
            purchase_serial,
            invoice.get("party_name"),
            invoice.get("place"),
            invoice.get("gstin"),
            invoice.get("invoice_no"),
            invoice_date,
            invoice.get("tax_rate"),
            invoice.get("hsn_code"),
            net_amount,
            cgst,
            sgst,
            igst,
            total_gst,
            gross_amount
        ]

        for col_num, value in enumerate(
            row_data,
            start=1
        ):

            cell = ws.cell(
                purchase_row,
                col_num
            )

            cell.value = value
            cell.border = thin_border
            cell.alignment = center_align

        purchase_row += 1
        purchase_serial += 1

    # PURCHASE TOTAL

    ws.cell(purchase_row, 1).value = "TOTAL"
    ws.cell(purchase_row, 1).font = header_font

    totals = [
        purchase_net_total,
        purchase_cgst_total,
        purchase_sgst_total,
        purchase_igst_total,
        purchase_total_gst,
        purchase_gross_total
    ]

    for idx, value in enumerate(
        totals,
        start=9
    ):

        cell = ws.cell(
            purchase_row,
            idx
        )

        cell.value = value
        cell.border = thin_border
        cell.alignment = center_align

    # ==========================
    # SALES SECTION
    # ==========================

    sales_start = purchase_row + 4

    ws.cell(sales_start, 1).value = "SALES"
    ws.cell(sales_start, 1).font = header_font

    for col_num, header in enumerate(headers, start=1):

        cell = ws.cell(
            sales_start + 1,
            col_num
        )

        cell.value = header
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align

    sales_row = sales_start + 2
    sales_serial = 1

    sales_net_total = 0
    sales_cgst_total = 0
    sales_sgst_total = 0
    sales_igst_total = 0
    sales_total_gst = 0
    sales_gross_total = 0

    for invoice in invoices:

        if invoice.get("invoice_type") != "Sales":
            continue

        try:
            invoice_date = datetime.strptime(
                invoice.get("invoice_date", ""),
                "%Y-%m-%d"
            ).strftime("%d-%m-%Y")
        except:
            invoice_date = invoice.get(
                "invoice_date",
                ""
            )

        net_amount = float(invoice.get("net_amount", 0) or 0)
        cgst = float(invoice.get("cgst", 0) or 0)
        sgst = float(invoice.get("sgst", 0) or 0)
        igst = float(invoice.get("igst", 0) or 0)

        total_gst = cgst + sgst + igst
        gross_amount = net_amount + total_gst

        sales_net_total += net_amount
        sales_cgst_total += cgst
        sales_sgst_total += sgst
        sales_igst_total += igst
        sales_total_gst += total_gst
        sales_gross_total += gross_amount

        row_data = [
            sales_serial,
            invoice.get("party_name"),
            invoice.get("place"),
            invoice.get("gstin"),
            invoice.get("invoice_no"),
            invoice_date,
            invoice.get("tax_rate"),
            invoice.get("hsn_code"),
            net_amount,
            cgst,
            sgst,
            igst,
            total_gst,
            gross_amount
        ]

        for col_num, value in enumerate(
            row_data,
            start=1
        ):

            cell = ws.cell(
                sales_row,
                col_num
            )

            cell.value = value
            cell.border = thin_border
            cell.alignment = center_align

        sales_row += 1
        sales_serial += 1

    # SALES TOTAL

    ws.cell(sales_row, 1).value = "TOTAL"
    ws.cell(sales_row, 1).font = header_font

    totals = [
        sales_net_total,
        sales_cgst_total,
        sales_sgst_total,
        sales_igst_total,
        sales_total_gst,
        sales_gross_total
    ]

    for idx, value in enumerate(
        totals,
        start=9
    ):

        cell = ws.cell(
            sales_row,
            idx
        )

        cell.value = value
        cell.border = thin_border
        cell.alignment = center_align

    # ==========================
    # GST SUMMARY
    # ==========================

    summary_row = sales_row + 4

    ws.cell(summary_row, 1).value = "GST SUMMARY"
    ws.cell(summary_row, 1).font = title_font

    ws.cell(summary_row + 2, 1).value = "PARTICULARS"
    ws.cell(summary_row + 2, 3).value = "CGST"
    ws.cell(summary_row + 2, 5).value = "SGST"
    ws.cell(summary_row + 2, 7).value = "IGST"

    ws.cell(summary_row + 3, 1).value = "INPUT"
    ws.cell(summary_row + 3, 3).value = purchase_cgst_total
    ws.cell(summary_row + 3, 5).value = purchase_sgst_total
    ws.cell(summary_row + 3, 7).value = purchase_igst_total

    ws.cell(summary_row + 4, 1).value = "OUTPUT"
    ws.cell(summary_row + 4, 3).value = sales_cgst_total
    ws.cell(summary_row + 4, 5).value = sales_sgst_total
    ws.cell(summary_row + 4, 7).value = sales_igst_total

    ws.cell(summary_row + 5, 1).value = "BALANCE"
    ws.cell(summary_row + 5, 3).value = sales_cgst_total - purchase_cgst_total
    ws.cell(summary_row + 5, 5).value = sales_sgst_total - purchase_sgst_total
    ws.cell(summary_row + 5, 7).value = sales_igst_total - purchase_igst_total

    # Freeze Header

    ws.freeze_panes = "A8"

    safe_name = (
        business_name
        .replace(" ", "_")
        .replace("/", "_")
    )

    output_file = f"{safe_name}_2026.xlsx"

    wb.save(output_file)

    return output_file
